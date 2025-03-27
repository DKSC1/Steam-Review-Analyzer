# file_handler.py
import os
import csv
import traceback
import re  # <--- Import re for regex
import shutil # <--- Import shutil for file operations
from tkinter import messagebox

# Try importing pandas, set flag
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    # Log warning during initialization maybe? Or let the function handle it.

# Assume utils.py and config.py are accessible
# Need get_game_folder_path from process_handler
from process_handler import get_game_folder_path


# --- XLSX Generation ---
# (generate_xlsx_from_csv stays the same)
def generate_xlsx_from_csv(csv_filepath, xlsx_filepath, log_func):
    """Reads a CSV and generates an XLSX file using pandas, if available."""
    if not PANDAS_AVAILABLE:
        log_func("XLSX generation skipped: 'pandas' or 'openpyxl' library not found.")
        log_func("Install using: pip install pandas openpyxl")
        return False

    if not os.path.exists(csv_filepath):
        log_func(f"XLSX generation skipped: Input CSV file not found: '{os.path.basename(csv_filepath)}'")
        return False

    log_func(f"Generating XLSX from '{os.path.basename(csv_filepath)}'...")
    try:
        # Read CSV using pandas
        df = pd.read_csv(csv_filepath)

        # Write to Excel using openpyxl engine
        with pd.ExcelWriter(xlsx_filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            # Auto-adjust column widths (within the ExcelWriter context)
            worksheet = writer.sheets['Sheet1']
            for col_idx, column in enumerate(df.columns):
                try:
                    # Calculate max length: header vs content
                    # Ensure content is treated as string for length calculation
                    max_len = max(
                        len(str(column)), # Header length
                        df[column].astype(str).map(len).max() # Max content length
                    )
                    # Add some padding, set min/max width
                    adjusted_width = min(max(max_len + 2, 10), 60) # Pad, Min 10, Max 60
                    # Get column letter (requires openpyxl utility or calculation)
                    # openpyxl is used by the engine, so it should be available
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_idx + 1)
                    worksheet.column_dimensions[col_letter].width = adjusted_width
                except Exception as col_e:
                    log_func(f"Warning: Could not auto-adjust width for column '{column}': {col_e}")
                    # Apply a default fallback width if adjustment fails
                    try:
                         from openpyxl.utils import get_column_letter
                         col_letter = get_column_letter(col_idx + 1)
                         worksheet.column_dimensions[col_letter].width = 20
                    except Exception:
                         pass # Ignore if even fallback fails

        log_func(f"Successfully generated XLSX file: '{os.path.basename(xlsx_filepath)}'.")
        return True

    except FileNotFoundError:
        # Should be caught by the initial os.path.exists check, but handle defensively
        log_func(f"Error generating XLSX: Input CSV '{os.path.basename(csv_filepath)}' disappeared.")
        messagebox.showerror("XLSX Generation Error", f"Could not find the CSV file needed to create the XLSX:\n{os.path.basename(csv_filepath)}") # noqa
        return False
    except ImportError:
        # Should be caught by PANDAS_AVAILABLE check, but handle defensively
        log_func("Error generating XLSX: pandas or openpyxl missing (ImportError).")
        messagebox.showerror("XLSX Generation Error", "Required libraries for XLSX support (pandas, openpyxl) are missing.\nInstall using: pip install pandas openpyxl") # noqa
        return False
    except Exception as e:
        log_func(f"Error generating XLSX file '{os.path.basename(xlsx_filepath)}': {e}")
        log_func(f"Traceback: {traceback.format_exc()}")
        messagebox.showerror("XLSX Generation Error", f"Failed to generate the XLSX file:\n{e}")
        return False


# --- Load Existing Data ---
# (load_existing_data stays the same)
def load_existing_data(widgets, log_func):
    """Loads existing processed data (CSV and AI Text) for the selected game."""
    game_name = ""
    steam_app_id = ""
    # Safely get widget values
    try:
        if widgets.get('game_name_entry'): game_name = widgets['game_name_entry'].get().strip()
        if widgets.get('steam_id_entry'): steam_app_id = widgets['steam_id_entry'].get().strip()
    except Exception as e:
        log_func(f"Error getting game/ID from widgets for loading: {e}")
        messagebox.showwarning("Input Error", "Could not read Game Name or Steam ID.")
        return False, None, None # Indicate failure

    if not game_name or not steam_app_id or not steam_app_id.isdigit():
        messagebox.showwarning("Input Error", "Game Name and valid Steam ID required to load existing data.")
        return False, None, None # Indicate failure

    # Get game folder path using the function from process_handler
    game_folder_path = get_game_folder_path(game_name, steam_app_id, log_func)
    if not game_folder_path:
        # Error already logged/shown by get_game_folder_path
        return False, None, None # Indicate failure

    log_func(f"Attempting to load existing processed data for {game_name}...")
    folder_basename = os.path.basename(game_folder_path)

    # Define expected file paths
    extracted_csv_path = os.path.join(game_folder_path, f"{folder_basename}_ai_extracted_data.csv")
    # Load the text file that EXCLUDES the CSV block (created by api_handler.send_to_ai)
    response_text_path = os.path.join(game_folder_path, f"{folder_basename}_ai_response_text.txt")

    loaded_csv_data = None
    loaded_text_data = None
    found_any_data = False

    # Try loading CSV Data
    if os.path.exists(extracted_csv_path):
        log_func(f"Found extracted data file: {os.path.basename(extracted_csv_path)}")
        try:
            with open(extracted_csv_path, 'r', encoding='utf-8', newline='') as f:
                reader = csv.reader(f)
                # Read and filter out completely empty rows
                loaded_csv_data = [row for row in reader if any(field.strip() for field in row)]
            if loaded_csv_data:
                log_func(f"Loaded {len(loaded_csv_data)} rows from extracted CSV file.")
                found_any_data = True
            else:
                log_func("Extracted CSV file was found but appears to be empty.")
                loaded_csv_data = None # Treat as no data if empty
        except (IOError, csv.Error) as e:
            log_func(f"Error loading extracted CSV data: {e}")
            messagebox.showwarning("Load Warning", f"Could not read the previously extracted CSV data file:\n{e}")
            loaded_csv_data = None # Mark as failed
        except Exception as e_csv_load:
             log_func(f"Unexpected error loading CSV: {e_csv_load}")
             messagebox.showwarning("Load Error", f"Unexpected error loading CSV:\n{e_csv_load}")
             loaded_csv_data = None
    else:
        log_func("No previously extracted CSV data file found.")

    # Try loading AI Response Text (the one excluding the CSV block)
    if os.path.exists(response_text_path):
        log_func(f"Found AI response text file: {os.path.basename(response_text_path)}")
        try:
            with open(response_text_path, 'r', encoding='utf-8') as f:
                loaded_text_data = f.read()
            log_func("Loaded AI response text data.")
            found_any_data = True
        except IOError as e:
            log_func(f"Error loading AI response text file: {e}")
            messagebox.showwarning("Load Warning", f"Could not read the previously saved AI response text:\n{e}")
            loaded_text_data = "[Error loading saved AI response text]" # Provide error text
        except Exception as e_text_load:
             log_func(f"Unexpected error loading text: {e_text_load}")
             messagebox.showwarning("Load Error", f"Unexpected error loading text:\n{e_text_load}")
             loaded_text_data = "[Error loading saved AI response text]"
    else:
        log_func("No previously saved AI response text file found.")

    # Check if any data was actually found and loaded successfully
    if not found_any_data:
        log_func("No existing processed data (CSV or Text) found for this game.")
        messagebox.showinfo("Load Data", "No previously generated AI data (Text or CSV) found for this game.")
        # Return False even if files existed but were empty or failed to load
        return False, None, None

    # Return success (meaning *some* data was loaded), loaded CSV (or None), loaded Text (or placeholder/None)
    return True, loaded_csv_data, loaded_text_data


# --- MODIFIED: Strip Metadata Function (Overwrites Optimized File) ---
def strip_review_metadata(widgets, log_func):
    """
    Reads the optimized review file, removes metadata headers (like 'Date YYYY-MM-DD Playtime...'),
    and overwrites the *original* optimized file with the stripped content.
    Uses a temporary file for safety.
    Returns True on success, False on failure.
    """
    game_name = ""
    steam_app_id = ""
    # Safely get widget values
    try:
        if widgets.get('game_name_entry'): game_name = widgets['game_name_entry'].get().strip()
        if widgets.get('steam_id_entry'): steam_app_id = widgets['steam_id_entry'].get().strip()
    except Exception as e:
        log_func(f"Error getting game/ID from widgets for stripping: {e}")
        messagebox.showwarning("Input Error", "Could not read Game Name or Steam ID.")
        return False

    if not game_name or not steam_app_id or not steam_app_id.isdigit():
        messagebox.showwarning("Input Error", "Game Name and valid Steam ID required to strip metadata.")
        return False

    # Get game folder path
    game_folder_path = get_game_folder_path(game_name, steam_app_id, log_func)
    if not game_folder_path:
        return False # Error already shown

    folder_basename = os.path.basename(game_folder_path)

    # Define the target file path (the optimized file)
    optimized_file_path = os.path.join(game_folder_path, f"{folder_basename}_reviews_optimized.txt")
    # Define a temporary file path in the same directory
    temp_file_path = os.path.join(game_folder_path, f"{folder_basename}_reviews_temp_strip.txt")

    # Check if optimized file exists
    if not os.path.exists(optimized_file_path):
        log_func(f"Stripping Error: Optimized file not found: {os.path.basename(optimized_file_path)}")
        messagebox.showwarning("File Missing", f"The optimized review file to modify was not found:\n{os.path.basename(optimized_file_path)}\n\nRun Step 2 (Optimize) first.")
        return False

    # --- Confirmation Prompt ---
    confirm_overwrite = messagebox.askyesno(
        "Confirm Overwrite",
        f"This will permanently remove metadata (like 'Date...', 'Playtime...', 'Rec...') "
        f"from the beginning of lines in the file:\n\n"
        f"'{os.path.basename(optimized_file_path)}'\n\n"
        f"This action cannot be undone easily. Proceed?",
        icon='warning'
    )
    if not confirm_overwrite:
        log_func("Metadata stripping cancelled by user (confirmation denied).")
        return False
    # --- End Confirmation ---

    log_func(f"Starting metadata stripping for: {os.path.basename(optimized_file_path)}")
    log_func("This will OVERWRITE the existing file.")

    # Define the improved regex pattern
    # - ^ : Start of the line
    # - Date\s+ : Literal 'Date' followed by one or more spaces
    # - \d{4}-\d{2}-\d{2} : YYYY-MM-DD format
    # - \s+Playtime\s+ : Space, 'Playtime', Space
    # - \d+h\s+\d+m : Digit(s) + "h", space(s), digit(s) + "m" format for playtime
    # - \s+Rec\s+ : Space, 'Rec', Space
    # - (?:Positive|Negative) : Non-capturing group for 'Positive' or 'Negative'
    # - \s* : Zero or more trailing spaces after the metadata
    metadata_pattern = re.compile(r'^Date\s+\d{4}-\d{2}-\d{2}\s+Playtime\s+\d+h\s+\d+m\s+Rec\s+(?:Positive|Negative)\s*')

    lines_processed = 0
    lines_stripped = 0
    success = False # Flag to track successful completion before replacing file

    try:
        # Process line by line to a temporary file first for safety
        with open(optimized_file_path, 'r', encoding='utf-8') as infile, \
             open(temp_file_path, 'w', encoding='utf-8') as temp_outfile:

            for line in infile:
                lines_processed += 1
                # Apply the regex substitution
                stripped_line = metadata_pattern.sub('', line)

                # Check if stripping actually removed something
                if len(stripped_line) < len(line):
                    lines_stripped += 1

                # Write the potentially modified line to the temporary output file
                temp_outfile.write(stripped_line)

        # If we reached here without exceptions, the temp file is complete
        success = True
        log_func(f"Successfully processed {lines_processed} lines to temporary file.")
        log_func(f"Removed metadata header from approximately {lines_stripped} lines.")

    except IOError as e:
        log_func(f"Stripping Error: Failed to read/write files: {e}")
        messagebox.showerror("File I/O Error", f"Failed to read or write files during stripping:\n{e}")
        success = False # Ensure replacement doesn't happen
    except Exception as e:
        log_func(f"Stripping Error: An unexpected error occurred during processing: {e}\n{traceback.format_exc()}")
        messagebox.showerror("Stripping Error", f"An unexpected error occurred during metadata stripping:\n{e}")
        success = False # Ensure replacement doesn't happen
    finally:
        # --- Overwrite original file ONLY if processing was successful ---
        if success:
            try:
                # Replace the original optimized file with the temporary stripped file
                shutil.move(temp_file_path, optimized_file_path)
                log_func(f"Successfully overwrote '{os.path.basename(optimized_file_path)}' with stripped content.")
                # Return True only after successful move/overwrite
                return True
            except OSError as move_err:
                log_func(f"CRITICAL ERROR: Failed to overwrite original file '{os.path.basename(optimized_file_path)}' after stripping: {move_err}")
                messagebox.showerror("File Replace Error", f"Failed to save the stripped content over the original file:\n{move_err}\n\nThe stripped content might be in a temporary file:\n{os.path.basename(temp_file_path)}")
                return False # Return False as the overwrite failed
            except Exception as final_err:
                log_func(f"CRITICAL ERROR: Unexpected error replacing file: {final_err}")
                messagebox.showerror("File Replace Error", f"Unexpected error saving stripped content:\n{final_err}")
                return False
        else:
            # If processing failed, attempt to clean up the temporary file
            if os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                    log_func("Removed temporary stripping file due to processing error.")
                except OSError as rm_err:
                    log_func(f"Warning: Could not remove temporary stripping file after error: {rm_err}")
            # Return False because the operation didn't complete successfully
            return False